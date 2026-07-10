"""Lead Finder App — Flask backend"""

import json
import logging
import os
import re
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from difflib import SequenceMatcher
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, url_for
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
from sqlalchemy.orm import defer as _sa_defer

from models import Lead, User, db

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("leadfinder")

# Ring buffer met recente logregels, uitleesbaar via /api/debug
import collections

_log_buffer = collections.deque(maxlen=300)


class _BufferLogHandler(logging.Handler):
    def emit(self, record):
        try:
            _log_buffer.append(self.format(record))
        except Exception:
            pass


_buffer_handler = _BufferLogHandler()
_buffer_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
logging.getLogger().addHandler(_buffer_handler)

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-change-in-production')

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

if not DATABASE_URL:
    pghost = os.environ.get('PGHOST', '')
    pgport = os.environ.get('PGPORT', '5432')
    pguser = os.environ.get('PGUSER', '')
    pgpassword = os.environ.get('PGPASSWORD', '')
    pgdatabase = os.environ.get('PGDATABASE', '')
    if pghost and pguser:
        DATABASE_URL = f"postgresql://{pguser}:{pgpassword}@{pghost}:{pgport}/{pgdatabase}"

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL or 'sqlite:///leads.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_page'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def _migrate_screenshots_to_column():
    """Eenmalig: verplaats screenshots uit de data-JSON naar de aparte kolom.
    Rij voor rij, zodat het geheugen laag blijft."""
    lead_ids = [r[0] for r in db.session.query(Lead.id).all()]
    migrated = 0
    for lead_id in lead_ids:
        row = db.session.get(Lead, lead_id)
        if row is None:
            continue
        data = row.data or {}
        shot = data.get("screenshot", "")
        if shot:
            row.screenshot = shot
            row.data = {k: v for k, v in data.items() if k != "screenshot"}
            migrated += 1
            db.session.commit()
        db.session.expunge_all()
    if migrated:
        log.info("Migratie: %d screenshots verplaatst naar aparte kolom", migrated)


with app.app_context():
    db.create_all()
    try:
        from sqlalchemy import inspect as _sa_inspect
        from sqlalchemy import text as _sa_text
        _cols = [c["name"] for c in _sa_inspect(db.engine).get_columns("leads")]
        if "screenshot" not in _cols:
            db.session.execute(_sa_text("ALTER TABLE leads ADD COLUMN screenshot TEXT"))
            db.session.commit()
        _migrate_screenshots_to_column()
    except Exception:
        log.exception("Screenshot-migratie mislukt")

SCREENSHOTS_DIR = os.path.join(os.path.dirname(__file__), "screenshots")
DATA_FILE = os.path.join(os.path.dirname(__file__), "leads_db.json")
EVENTS_FILE = os.path.join(os.path.dirname(__file__), "events_db.json")
EMAIL_TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "static", "email-templates")
EMAILS_DIR = os.path.join(os.path.dirname(__file__), "emails")
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
os.makedirs(EMAILS_DIR, exist_ok=True)

def _load_template(name: str) -> str:
    path = os.path.join(EMAIL_TEMPLATES_DIR, f"{name}.html")
    try:
        with open(path, encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return ""

EMAIL_TEMPLATES = {
    "voetbal": _load_template("voetbal"),
    "tennis": _load_template("tennis"),
    "hockey": _load_template("hockey"),
    "golf": _load_template("golf"),
    "eten": _load_template("eten"),
}

jobs = {}
_leads_lock = threading.Lock()
_job_executor = ThreadPoolExecutor(max_workers=4)

# Max 3 browsers tegelijk op de héle server. Elke Chromium kost ~400MB.
# Historie: t/m 2026-07-09 stond dit op 1 omdat het Railway-trialplan maar 1GB
# geheugen gaf (2 gelijktijdige launches → 920MB+ en alles bevroor). Op 2026-07-10
# is geüpgraded naar het Hobby-plan (geheugengrens 48GB), dus 3 kan nu veilig.
_BROWSER_SLOTS_TOTAL = 3
_browser_slots = threading.Semaphore(_BROWSER_SLOTS_TOTAL)
_slot_holders: dict = {}  # thread-id -> (omschrijving, starttijd)
_slot_lock = threading.Lock()
_SLOT_STUCK_SECONDS = 240


def _acquire_browser_slot(what: str, timeout_s: int = 120) -> bool:
    if _browser_slots.acquire(timeout=timeout_s):
        with _slot_lock:
            _slot_holders[threading.get_ident()] = (what, time.time())
        return True
    log.error("Geen browser-slot vrij na %ds voor %s — server te druk", timeout_s, what)
    return False


def _release_browser_slot():
    with _slot_lock:
        _slot_holders.pop(threading.get_ident(), None)
    _browser_slots.release()


def _wait_until_slot_free(job_id: str, max_wait_s: int = 600):
    """Als alle browser-slots bezet zijn: eerlijk melden dat het druk is en wachten.

    Voorkomt dat een gebruiker na de slot-timeout stilletjes "Geen leads gevonden"
    te zien krijgt terwijl de server gewoon bezig was voor iemand anders."""
    deadline = time.time() + max_wait_s
    logged = False
    while time.time() < deadline:
        with _slot_lock:
            slot_free = len(_slot_holders) < _BROWSER_SLOTS_TOTAL
        if slot_free:
            return
        job = jobs.get(job_id)
        if job is None or job.get("status") != "running":
            return
        if not logged:
            log.info("Job %s wacht: alle browser-slots bezet", job_id)
            logged = True
        job["message"] = "Druk op de server: een andere zoekopdracht draait nu — de jouwe start automatisch zodra die klaar is..."
        time.sleep(5)


def _is_zombie(pid: str) -> bool:
    try:
        with open(f"/proc/{pid}/stat") as f:
            # Formaat: "pid (comm) state ..." — comm kan spaties bevatten
            return f.read().rsplit(")", 1)[1].split()[0] == "Z"
    except Exception:
        return False


def _list_chromium_pids(include_zombies: bool = False) -> list:
    pids = []
    try:
        for pid in os.listdir("/proc"):
            if not pid.isdigit():
                continue
            try:
                with open(f"/proc/{pid}/comm") as f:
                    name = f.read().strip().lower()
                if ("chrom" in name or "headless" in name) and (include_zombies or not _is_zombie(pid)):
                    pids.append(int(pid))
            except Exception:
                continue
    except Exception:
        pass
    return pids


def _reap_zombies() -> int:
    """Ruim zombie-kindprocessen op (python is PID 1 in de container en erft alle wezen)."""
    reaped = 0
    while True:
        try:
            pid, _ = os.waitpid(-1, os.WNOHANG)
        except (ChildProcessError, OSError):
            break
        if pid == 0:
            break
        reaped += 1
    return reaped


def _zombie_sweeper():
    """Ruimt achtergebleven of vastgelopen chromium-processen op, zodat het geheugen nooit dichtslibt."""
    while True:
        time.sleep(120)
        try:
            now = time.time()
            with _slot_lock:
                holders = list(_slot_holders.values())
            stuck = [what for what, started in holders if now - started > _SLOT_STUCK_SECONDS]
            idle = not holders
            if not idle and not stuck:
                continue
            pids = _list_chromium_pids()
            if pids:
                if stuck:
                    log.warning("Sweeper: vastgelopen browsertaken %s — %d chromium-processen hard opruimen", stuck, len(pids))
                else:
                    log.warning("Sweeper: %d achtergebleven chromium-processen zonder actieve taak — opruimen", len(pids))
                for pid in pids:
                    try:
                        os.kill(pid, 9)
                    except Exception:
                        pass
            if idle:
                reaped = _reap_zombies()
                if reaped:
                    log.info("Sweeper: %d zombie-processen definitief opgeruimd", reaped)
        except Exception:
            log.exception("Sweeper-fout")


if os.name == "posix":
    threading.Thread(target=_zombie_sweeper, daemon=True).start()

# Zuinige browserinstellingen — de server heeft maar 1GB geheugen
_CHROMIUM_ARGS = [
    "--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage",
    "--disable-gpu", "--no-zygote", "--disable-extensions",
    "--renderer-process-limit=2", "--disable-background-networking",
]

SPORT_KEYWORDS = ["voetbal", "hockey", "padel", "tennis", "basketbal", "volleybal",
                  "handbal", "zwemclub", "atletiek", "rugby", "cricket", "badminton"]

BOARD_ROLES = ["voorzitter", "secretaris", "penningmeester", "kantinecommissie"]


# ── Data ──────────────────────────────────────────────────────────────────────

def _email_file(lead_id: str) -> str:
    return os.path.join(EMAILS_DIR, f"{lead_id}.html")


def save_email(lead_id: str, html: str):
    with open(_email_file(lead_id), "w", encoding="utf-8") as f:
        f.write(html)


def load_email(lead_id: str) -> str:
    path = _email_file(lead_id)
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return ""


def delete_email(lead_id: str):
    path = _email_file(lead_id)
    if os.path.exists(path):
        os.remove(path)


def _load_leads_unsafe(user_id=None):
    if user_id is None:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return []
    with app.app_context():
        # Alleen data + heeft-screenshot-vlag selecteren. NOOIT hele Lead-rijen
        # laden: dan haalt SQLAlchemy ook de screenshot-kolom (~0,5MB/lead) op
        # en loopt het geheugen vol zodra er veel leads zijn.
        rows = (
            db.session.query(Lead.data, Lead.screenshot.isnot(None))
            .filter(Lead.user_id == user_id)
            .order_by(Lead.created_at.desc())
            .all()
        )
        result = []
        for data, has_shot in rows:
            d = dict(data)
            d["has_screenshot"] = bool(has_shot)
            result.append(d)
        return result


def _save_leads_unsafe(leads, user_id=None):
    if user_id is None:
        slim = [{k: v for k, v in l.items() if k != "cold_email"} for l in leads]
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(slim, f, ensure_ascii=False, indent=2)
        return
    with app.app_context():
        # defer(screenshot): rijen laden zonder de zware screenshot-kolom
        existing_ids = {
            row.id: row
            for row in Lead.query.options(_sa_defer(Lead.screenshot)).filter_by(user_id=user_id).all()
        }
        new_ids = {l["id"] for l in leads}
        for lead_id in list(existing_ids.keys()):
            if lead_id not in new_ids:
                db.session.delete(existing_ids[lead_id])
        for lead in leads:
            # Screenshot gaat naar de aparte kolom, nooit in de data-JSON
            slim = {k: v for k, v in lead.items() if k not in ("cold_email", "screenshot", "has_screenshot")}
            shot = lead.get("screenshot", "")
            if lead["id"] in existing_ids:
                row = existing_ids[lead["id"]]
                row.data = slim
                if shot:
                    row.screenshot = shot
            else:
                db.session.add(Lead(id=lead["id"], user_id=user_id, data=slim, screenshot=shot or None))
        db.session.commit()


def load_leads(user_id=None):
    with _leads_lock:
        leads = _load_leads_unsafe(user_id)

    changed = False
    for lead in leads:
        if lead.get("cold_email"):
            try:
                save_email(lead["id"], lead["cold_email"])
                lead["cold_email"] = ""
                changed = True
            except Exception:
                pass

    if changed:
        with _leads_lock:
            _save_leads_unsafe(leads, user_id)

    return leads


def save_leads(leads, user_id=None):
    with _leads_lock:
        _save_leads_unsafe(leads, user_id)


def add_lead(lead: dict, user_id=None) -> bool:
    """Voeg een lead toe. Returns False als het een duplicaat is."""
    with _leads_lock:
        leads = _load_leads_unsafe(user_id)
        lead_city = lead.get("city", "").lower().strip()
        for existing in leads:
            if existing.get("city", "").lower().strip() == lead_city:
                if _names_similar(lead["name"], existing["name"]):
                    return False
        leads.insert(0, lead)
        _save_leads_unsafe(leads, user_id)
        return True


def load_events():
    if os.path.exists(EVENTS_FILE):
        with open(EVENTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_events(events):
    with open(EVENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)


def is_sport(niche: str) -> bool:
    return any(k in niche.lower() for k in SPORT_KEYWORDS) or "club" in niche.lower() or "vereniging" in niche.lower()


_NL_CITIES = [
    ("Amsterdam", 52.3676, 4.9041), ("Rotterdam", 51.9225, 4.4792), ("Den Haag", 52.0705, 4.3007),
    ("Utrecht", 52.0908, 5.1222), ("Eindhoven", 51.4416, 5.4697), ("Tilburg", 51.5555, 5.0913),
    ("Groningen", 53.2194, 6.5665), ("Almere", 52.3508, 5.2647), ("Breda", 51.5719, 4.7683),
    ("Nijmegen", 51.8426, 5.8546), ("Enschede", 52.2215, 6.8937), ("Haarlem", 52.3874, 4.6462),
    ("Arnhem", 51.9851, 5.8987), ("Zaanstad", 52.4563, 4.8180), ("Amersfoort", 52.1561, 5.3878),
    ("Apeldoorn", 52.2112, 5.9699), ("Den Bosch", 51.6978, 5.3037), ("Hoofddorp", 52.3025, 4.6939),
    ("Maastricht", 50.8514, 5.6910), ("Leiden", 52.1601, 4.4970), ("Dordrecht", 51.8133, 4.6901),
    ("Zoetermeer", 52.0577, 4.4940), ("Zwolle", 52.5168, 6.0830), ("Deventer", 52.2550, 6.1552),
    ("Delft", 52.0116, 4.3571), ("Alkmaar", 52.6324, 4.7534), ("Leeuwarden", 53.2012, 5.7999),
    ("Westland", 51.9959, 4.2228), ("Emmen", 52.7791, 6.8990), ("Venlo", 51.3704, 6.1724),
    ("Ede", 52.0438, 5.6647), ("Lelystad", 52.5185, 5.4714), ("Sittard", 51.0025, 5.8702),
    ("Hilversum", 52.2292, 5.1772), ("Purmerend", 52.5037, 4.9600), ("Roosendaal", 51.5311, 4.4614),
    ("Spijkenisse", 51.8445, 4.3289), ("Schiedam", 51.9179, 4.3989), ("Heerlen", 50.8880, 5.9796),
    ("Helmond", 51.4814, 5.6580), ("Leidschendam", 52.0854, 4.3951), ("Alphen aan den Rijn", 52.1319, 4.6618),
    ("Amstelveen", 52.3098, 4.8598), ("Vlaardingen", 51.9124, 4.3428), ("Capelle aan den IJssel", 51.9330, 4.5723),
    ("Gouda", 52.0116, 4.7114), ("Oss", 51.7638, 5.5170), ("Nissewaard", 51.8320, 4.2870),
    ("Bergen op Zoom", 51.4949, 4.2888), ("Dronten", 52.5246, 5.7195), ("Zeist", 52.0889, 5.2336),
    ("Hoorn", 52.6440, 5.0608), ("Middelburg", 51.4988, 3.6136), ("Hengelo", 52.2659, 6.7936),
    ("Almelo", 52.3568, 6.6626), ("Doetinchem", 51.9661, 6.2955), ("Harderwijk", 52.3452, 5.6224),
    ("Woerden", 52.0886, 4.8834), ("Weert", 51.2525, 5.7069), ("Zaandam", 52.4386, 4.8327),
    ("Veenendaal", 52.0269, 5.5557), ("Nieuwegein", 52.0343, 5.0841), ("Barneveld", 52.1418, 5.5876),
    ("Roermond", 51.1933, 5.9893), ("Kerkrade", 50.8659, 6.0679), ("Terneuzen", 51.3353, 3.8299),
    ("Hoogeveen", 52.7279, 6.4769), ("Stadskanaal", 52.9863, 6.9509), ("Meppel", 52.6963, 6.1945),
    ("Assen", 52.9929, 6.5642), ("Emmeloord", 52.7134, 5.7501), ("Drachten", 53.1046, 6.0968),
    ("Sneek", 53.0328, 5.6574), ("Heerenveen", 52.9607, 5.9257), ("Franeker", 53.1868, 5.5424),
    ("Vlissingen", 51.4561, 3.5722), ("Goes", 51.5044, 3.8897), ("Zierikzee", 51.6500, 3.9167),
    ("Delfzijl", 53.3290, 6.9236), ("Winschoten", 53.1437, 7.0400), ("Coevorden", 52.6647, 6.7443),
    ("Hardenberg", 52.5758, 6.6175), ("Oldenzaal", 52.3127, 6.9289), ("Losser", 52.2611, 7.0050),
    ("Waalwijk", 51.6885, 5.0693), ("Dongen", 51.6266, 4.9372), ("Geertruidenberg", 51.7007, 4.8605),
    ("Gorinchem", 51.8338, 4.9752), ("Papendrecht", 51.8297, 4.6948), ("Ridderkerk", 51.8671, 4.6067),
    ("Barendrecht", 51.8576, 4.5353), ("Lansingerland", 51.9700, 4.5900), ("Pijnacker", 52.0057, 4.4303),
    ("Naaldwijk", 51.9970, 4.2127), ("Monster", 51.9890, 4.1716), ("Midden-Delfland", 51.9814, 4.3145),
    ("Wassenaar", 52.1459, 4.4011), ("Voorburg", 52.0741, 4.3644), ("Rijswijk", 52.0415, 4.3221),
]

def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    import math
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def get_nearby_places(city: str, radius_km: int) -> list:
    """Zoek plaatsen (inclusief dorpen) binnen radius_km kilometer van city via Overpass API."""
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": city + ", Nederland", "format": "json", "limit": 1},
            headers={"User-Agent": "HeinekenLeadFinder/1.0"},
            timeout=5
        )
        data = r.json()
        if not data:
            return []
        lat = float(data[0]["lat"])
        lon = float(data[0]["lon"])

        radius_m = radius_km * 1000
        query = f"""
[out:json][timeout:10];
(
  node["place"~"city|town|village|hamlet"](around:{radius_m},{lat},{lon});
);
out;
"""
        resp = requests.post(
            "https://overpass-api.de/api/interpreter",
            data=query,
            headers={"User-Agent": "HeinekenLeadFinder/1.0"},
            timeout=12
        )
        elements = resp.json().get("elements", [])
        seen = {city.lower()}
        nearby = []
        for el in elements:
            name = el.get("tags", {}).get("name:nl") or el.get("tags", {}).get("name", "")
            if not name or name.lower() in seen:
                continue
            seen.add(name.lower())
            elat = el.get("lat", lat)
            elon = el.get("lon", lon)
            dist = _haversine(lat, lon, elat, elon)
            nearby.append((name, dist))

        nearby.sort(key=lambda x: x[1])
        return [n for n, _ in nearby[:10]]
    except Exception:
        # Fallback op hardcoded lijst als Overpass niet bereikbaar is
        city_lower = city.lower()
        fallback = []
        for name, clat, clon in _NL_CITIES:
            if name.lower() == city_lower:
                continue
            if _haversine(lat, lon, clat, clon) <= radius_km:
                fallback.append((name, _haversine(lat, lon, clat, clon)))
        fallback.sort(key=lambda x: x[1])
        return [n for n, _ in fallback[:10]]


def get_gemeente(city: str) -> str:
    """Zoek de gemeente op van een dorp via OpenStreetMap Nominatim."""
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": city + ", Nederland", "format": "json", "addressdetails": 1, "limit": 1},
            headers={"User-Agent": "HeinekenLeadFinder/1.0"},
            timeout=5
        )
        data = r.json()
        if data:
            addr = data[0].get("address", {})
            gemeente = addr.get("municipality") or addr.get("city") or addr.get("town") or addr.get("village") or ""
            if gemeente.lower() != city.lower():
                return gemeente
    except Exception:
        pass
    return ""


# ── Google Maps scraper ───────────────────────────────────────────────────────

def scrape_google_maps(niche: str, city: str, max_results: int, results: list = None) -> list:
    query = f"{niche} {city}".strip() if city and city.lower() not in ("", "nederland") else niche
    # De aanroeper (watchdog) kan een lijst meegeven zodat deelresultaten
    # niet verloren gaan als de scrape halverwege wordt afgebroken
    if results is None:
        results = []
    # Harde limiet zodat één vastgelopen scrape nooit de hele job blokkeert
    deadline = time.time() + 90
    log.info("Maps-scrape start: %r (max %d)", query, max_results)

    if not _acquire_browser_slot(f"maps:{query}"):
        return results
    try:
        return _scrape_google_maps_inner(query, niche, city, max_results, deadline, results)
    finally:
        _release_browser_slot()


def _scrape_google_maps_inner(query: str, niche: str, city: str, max_results: int, deadline: float, results: list) -> list:
    with sync_playwright() as p:
        browser = None
        try:
            browser = p.chromium.launch(headless=True, args=_CHROMIUM_ARGS)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800},
            )
            # Geen enkele browseractie mag oneindig wachten
            context.set_default_timeout(10000)
            page = context.new_page()
            try:
                page.goto(
                    f"https://www.google.com/maps/search/{query.replace(' ', '+')}",
                    wait_until="domcontentloaded", timeout=30000
                )
            except Exception as e:
                log.warning("Maps-scrape: pagina laden mislukt voor %r: %s", query, e)
                return results
            page.wait_for_timeout(1200)

            for text in ["Alles accepteren", "Accept all", "Akkoord"]:
                try:
                    page.click(f'button:has-text("{text}")', timeout=1500)
                    page.wait_for_timeout(300)
                    break
                except Exception:
                    pass

            page.wait_for_timeout(600)

            if "consent.google.com" in page.url or "sorry" in page.url:
                log.warning("Maps-scrape: geblokkeerd door Google (url=%s)", page.url)
                return results

            place_urls = []
            seen_names = set()
            no_new_count = 0

            # Wacht tot óf de resultatenlijst óf het detailpaneel geladen is
            try:
                page.wait_for_selector(
                    'a[href*="/maps/place/"], [data-item-id="address"], a[data-item-id="authority"]',
                    timeout=8000,
                )
            except Exception:
                pass

            # Bij een exacte match toont Google direct de detailpagina (soms via
            # redirect naar /maps/place/, soms op de zoek-URL zelf) zonder
            # resultatenlijst — behandel die pagina dan als het enige resultaat
            direct_match = "/maps/place/" in page.url
            if not direct_match:
                try:
                    if page.query_selector('[data-item-id="address"]') or page.query_selector('a[data-item-id="authority"]'):
                        direct_match = True
                except Exception:
                    pass
            if direct_match:
                place_name = ""
                try:
                    el = page.query_selector("h1")
                    if el:
                        place_name = (el.inner_text() or "").strip()
                except Exception:
                    pass
                place_urls.append((place_name or niche, page.url))
                log.info("Maps-scrape: directe match op plaats-pagina: %r", place_name or niche)

            while not direct_match and len(place_urls) < max_results and time.time() < deadline:
                cards = page.query_selector_all('a[href*="/maps/place/"]')
                before = len(place_urls)
                for card in cards:
                    if len(place_urls) >= max_results:
                        break
                    name = (card.get_attribute("aria-label") or "").strip()
                    href = (card.get_attribute("href") or "").strip()
                    if name and href and name not in seen_names:
                        seen_names.add(name)
                        place_urls.append((name, href))
                if len(place_urls) == before:
                    no_new_count += 1
                    if no_new_count >= 3:
                        break
                else:
                    no_new_count = 0
                feed = page.query_selector('div[role="feed"]')
                if feed:
                    # Geen evaluate() gebruiken: die heeft géén timeout en kan
                    # eeuwig hangen als de pagina crasht — muiswiel is wel begrensd
                    try:
                        feed.hover(timeout=3000)
                        page.mouse.wheel(0, 800)
                    except Exception:
                        break
                    page.wait_for_timeout(600)
                else:
                    break

            if not place_urls:
                log.warning("Maps-scrape: 0 resultaatkaarten voor %r (url=%s)", query, page.url)

            for name, href in place_urls:
                if time.time() > deadline:
                    log.warning("Maps-scrape: deadline bereikt na %d/%d plaatsen voor %r", len(results), len(place_urls), query)
                    break
                try:
                    page.goto(href, wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_timeout(800)

                    address, phone, website, rating = "", "", "", ""

                    try:
                        el = page.query_selector('[data-item-id="address"] .fontBodyMedium')
                        if el:
                            address = el.inner_text().strip()
                    except Exception:
                        pass

                    try:
                        els = page.query_selector_all('[data-item-id^="phone"]')
                        for el in els:
                            t = el.inner_text().strip()
                            if t:
                                phone = t
                                break
                    except Exception:
                        pass

                    try:
                        el = page.query_selector('a[data-item-id="authority"]')
                        if el:
                            website = el.get_attribute("href") or el.inner_text().strip()
                        else:
                            el = page.query_selector('[data-item-id="authority"]')
                            if el:
                                website = el.inner_text().strip()
                    except Exception:
                        pass

                    try:
                        el = page.query_selector('span[aria-hidden="true"].fontDisplayLarge')
                        if el:
                            rating = el.inner_text().strip()
                    except Exception:
                        pass

                    results.append({
                        "id": str(uuid.uuid4()),
                        "type": "sport" if is_sport(niche) else "business",
                        "name": name,
                        "niche": niche,
                        "city": city,
                        "address": address,
                        "phone": phone,
                        "website": website,
                        "rating": rating,
                        "website_status": "",
                        "screenshot": "",
                        "email": "",
                        "contact_person": "",
                        "board": {},
                        "cold_email": "",
                        "found_at": datetime.now().isoformat(),
                        "status": "nieuw",
                    })
                except Exception as e:
                    log.warning("Maps-scrape: detailpagina mislukt voor %r: %s", name, e)
                    continue
        except Exception:
            log.exception("Maps-scrape: onverwachte fout voor %r", query)
        finally:
            if browser:
                try:
                    browser.close()
                except Exception:
                    pass

    log.info("Maps-scrape klaar: %r -> %d resultaten", query, len(results))
    return results


# ── Website checker ───────────────────────────────────────────────────────────

def check_website(website: str) -> str:
    if not website:
        return "missing"
    url = website if website.startswith("http") else f"https://{website}"
    try:
        r = requests.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code >= 400:
            return "bad"
        soup = BeautifulSoup(r.text, "html.parser")
        text = soup.get_text()
        word_count = len(text.split())
        has_contact = any(w in text.lower() for w in ["contact", "email", "phone", "tel"])
        is_mobile = "viewport" in r.text.lower()
        if word_count < 100 or not has_contact or not is_mobile:
            return "bad"
        return "good"
    except Exception:
        return "bad"


# ── Cookie banner dismissal ───────────────────────────────────────────────────

def _dismiss_cookie_banner(page) -> None:
    """Try to click common cookie accept buttons so they don't block screenshots."""
    # Meest specifieke teksten eerst — de "alles accepteren"-knop wint van losse toggles
    accept_texts = [
        "alle cookies accepteren", "accepteer alle cookies", "alles accepteren",
        "alle accepteren", "accept all cookies", "accept all", "allow all",
        "cookies toestaan", "allow cookies", "ja, ik ga akkoord", "ik ga akkoord",
        "accepteren", "accepteer", "akkoord", "toestaan", "accept", "agree", "ok",
    ]
    # Knoppen die we juist NIET willen raken ("niet toestaan", "weigeren", "instellingen")
    reject_words = ["niet", "weiger", "afwijz", "reject", "deny", "decline",
                    "instelling", "voorkeur", "settings", "aanpassen", "beheer",
                    "manage", "meer info", "lees meer"]

    def _click_accept_in(frame) -> bool:
        try:
            buttons = frame.query_selector_all(
                "button, a[role='button'], input[type='button'], input[type='submit'], [role='button']"
            )
        except Exception:
            return False
        best_btn, best_rank = None, len(accept_texts)
        for btn in buttons:
            try:
                text = (btn.inner_text() or btn.get_attribute("value") or "").strip().lower()
            except Exception:
                continue
            if not text or len(text) > 40:
                continue
            if any(w in text for w in reject_words):
                continue
            for rank, t in enumerate(accept_texts):
                if t in text:
                    if rank < best_rank:
                        best_btn, best_rank = btn, rank
                    break
        if best_btn:
            try:
                best_btn.click(timeout=1500)
                return True
            except Exception:
                return False
        return False

    try:
        # Twee rondes: sommige sites tonen na de eerste klik nog een tweede laag,
        # en banners zitten regelmatig in een iframe
        for _ in range(2):
            clicked = False
            for frame in page.frames:
                if _click_accept_in(frame):
                    clicked = True
                    page.wait_for_timeout(800)
                    break
            if not clicked:
                break
    except Exception:
        pass


# ── Screenshot ────────────────────────────────────────────────────────────────

def take_screenshot(website: str, lead_id: str) -> str:
    import base64
    if not website:
        return ""
    url = website if website.startswith("http") else f"https://{website}"
    if not _acquire_browser_slot(f"screenshot:{url}"):
        return ""
    try:
        with Stealth().use_sync(sync_playwright()) as p:
            browser = p.chromium.launch(headless=True, args=_CHROMIUM_ARGS + [
                "--disable-blink-features=AutomationControlled",
            ])
            try:
                context = browser.new_context(
                    viewport={"width": 1280, "height": 800},
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                    locale="nl-NL",
                )
                context.set_default_timeout(10000)
                page = context.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=15000)
                page.wait_for_timeout(1000)
                _dismiss_cookie_banner(page)
                page.wait_for_timeout(500)
                img_bytes = page.screenshot(full_page=False)
            finally:
                try:
                    browser.close()
                except Exception:
                    pass
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        return f"data:image/png;base64,{b64}"
    except Exception as e:
        log.warning("Screenshot mislukt voor %s: %s", url, e)
        return ""
    finally:
        _release_browser_slot()


# ── Logo finder ──────────────────────────────────────────────────────────────

def _fetch_html_with_playwright(url: str) -> str:
    """Haal de HTML op via een echte browser — voor sites die requests blokkeren (403/bot-detectie)."""
    if not _acquire_browser_slot(f"html:{url}"):
        return ""
    try:
        with Stealth().use_sync(sync_playwright()) as p:
            browser = p.chromium.launch(headless=True, args=_CHROMIUM_ARGS)
            try:
                context = browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                    locale="nl-NL",
                )
                context.set_default_timeout(10000)
                page = context.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=15000)
                page.wait_for_timeout(800)
                return page.content()
            finally:
                try:
                    browser.close()
                except Exception:
                    pass
    except Exception as e:
        log.warning("Playwright HTML-fallback mislukt voor %s: %s", url, e)
        return ""
    finally:
        _release_browser_slot()


def find_logo(website: str) -> str:
    """Find the club/business logo URL. Prefers favicons and small dedicated logo images."""
    if not website:
        return ""
    url = website if website.startswith("http") else f"https://{website}"
    base = url.rstrip("/")
    try:
        html = ""
        try:
            r = requests.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code < 400:
                html = r.text
        except Exception:
            pass
        # Sites met botbescherming (403) laden wel in een echte browser
        if not html:
            html = _fetch_html_with_playwright(url)
        if not html:
            log.warning("Logo zoeken: geen HTML voor %s", url)
            return ""
        soup = BeautifulSoup(html, "html.parser")

        def abs_url(src):
            if not src:
                return ""
            if src.startswith("http"):
                return src
            if src.startswith("//"):
                return "https:" + src
            return base + "/" + src.lstrip("/")

        def _icon_size(tag) -> int:
            """Pixelmaat van een icoon-link; onbekend = 180 (Apple-standaard)."""
            sizes = tag.get("sizes", "")
            if isinstance(sizes, (list, tuple)):
                sizes = " ".join(sizes)
            m = re.search(r"(\d+)x\d+", sizes) or re.search(r"(\d+)x\d+", tag.get("href", ""))
            return int(m.group(1)) if m else 180

        # 1. Grootste apple touch icon (kleine 57x57-varianten worden wazig in de e-mail)
        touch_icons = []
        for rel in ["apple-touch-icon", "apple-touch-icon-precomposed"]:
            for tag in soup.find_all("link", rel=rel):
                if tag.get("href"):
                    touch_icons.append((_icon_size(tag), abs_url(tag["href"])))
        touch_icons.sort(key=lambda t: t[0], reverse=True)
        if touch_icons and touch_icons[0][0] >= 120:
            return touch_icons[0][1]

        # 2. <img> with "logo" clearly in src, class or alt — skip large/wide images
        for img in soup.find_all("img"):
            src = img.get("src", "")
            alt = img.get("alt", "")
            classes = " ".join(img.get("class", []))
            attrs = f"{src} {alt} {classes}".lower()
            if "logo" not in attrs:
                continue
            # Skip obvious non-logos: photos, banners, headers, slideshows
            skip_words = ["banner", "header", "slide", "hero", "achtergrond", "background",
                          "nieuws", "news", "photo", "foto", "carousel", "sponsor"]
            if any(w in attrs for w in skip_words):
                continue
            if src:
                return abs_url(src)

        # 2b. Geen echt logo gevonden — dan toch de (kleinere) touch icon
        if touch_icons:
            return touch_icons[0][1]

        # 3. Favicon (png/svg/ico via <link> tags) — grootste eerst
        favicons = []
        for tag in soup.find_all("link"):
            rel = " ".join(tag.get("rel", [])).lower()
            if "icon" not in rel:
                continue
            href = tag.get("href", "")
            if not href:
                continue
            if any(href.split("?")[0].endswith(ext) for ext in (".png", ".svg", ".ico")):
                favicons.append((_icon_size(tag), abs_url(href)))
        if favicons:
            favicons.sort(key=lambda t: t[0], reverse=True)
            return favicons[0][1]

        # 4. Common logo paths
        for path in ["/logo.png", "/logo.svg", "/images/logo.png", "/img/logo.png",
                     "/assets/logo.png", "/wp-content/uploads/logo.png"]:
            try:
                resp = requests.head(base + path, timeout=3)
                if resp.status_code == 200:
                    return base + path
            except Exception:
                pass

        # 5. Default favicon.ico
        favicon = base + "/favicon.ico"
        try:
            resp = requests.head(favicon, timeout=3)
            if resp.status_code == 200:
                return favicon
        except Exception:
            pass

    except Exception:
        pass
    return ""


# ── Email + contact finder ────────────────────────────────────────────────────

def find_address_from_website(website: str) -> str:
    """Zoek adres op de website via Nederlandse postcode+plaatsnaam patronen."""
    if not website:
        return ""
    url = website if website.startswith("http") else f"https://{website}"
    base = url.rstrip("/")
    pages_to_check = [url, base + "/contact", base + "/contactpagina", base + "/over-ons", base + "/locatie"]
    # Volledig adres: straat + huisnummer + postcode (4 cijfers + 2 hoofdletters) + plaats
    adres_pattern = re.compile(
        r"[A-Z][a-záéíóúàèìòùäëïöü][a-záéíóúàèìòùäëïöü\s]{2,30}\s+\d{1,4}[a-z]?[,\s]+[1-9]\d{3}\s?[A-Z]{2}[\s,]+[A-Za-z][a-zA-Z\s\-]{2,20}",
        re.UNICODE
    )
    visited = set()
    for page_url in pages_to_check:
        if page_url in visited:
            continue
        visited.add(page_url)
        try:
            r = requests.get(page_url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if r.status_code >= 400:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            text = soup.get_text(" ", strip=True)
            m = adres_pattern.search(text)
            if m:
                raw = m.group(0).strip()
                # Stop bij woorden die geen adres meer zijn
                raw = re.split(r"(?:dit|route|klik|tel|email|meer|bekijk)", raw, flags=re.IGNORECASE)[0]
                raw = re.sub(r"\s{2,}", " ", raw)
                return raw.strip().rstrip(",")
        except Exception:
            continue
    return ""


def find_phone_from_website(website: str) -> str:
    """Zoek telefoonnummer op de website via tel: links en Nederlandse nummerpatronen."""
    if not website:
        return ""
    url = website if website.startswith("http") else f"https://{website}"
    base = url.rstrip("/")
    pages_to_check = [url, base + "/contact", base + "/contactpagina", base + "/over-ons"]
    phone_pattern = re.compile(
        r"(?:tel:\s*)?(\+31[\s\-]?(?:0[\s\-]?)?[1-9][\d\s\-]{7,11}|0[1-9][\d\s\-]{8,10}|06[\s\-]?[\d\s\-]{8})"
    )
    visited = set()
    for page_url in pages_to_check:
        if page_url in visited:
            continue
        visited.add(page_url)
        try:
            r = requests.get(page_url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if r.status_code >= 400:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            # Eerst tel: links checken — meest betrouwbaar
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("tel:"):
                    number = re.sub(r"[\s\-\(\)]", "", href[4:].strip())
                    if len(number) >= 10:
                        if number.startswith("0"):
                            number = "+31" + number[1:]
                        return number
            # Dan regex op de paginatekst
            text = soup.get_text()
            match = phone_pattern.search(text)
            if match:
                number = re.sub(r"[\s\-]", "", match.group(1).strip())
                if number.startswith("0"):
                    number = "+31" + number[1:]
                return number
        except Exception:
            continue
    return ""


def find_email_and_contact(website: str) -> tuple:
    if not website:
        return "", ""
    url = website if website.startswith("http") else f"https://{website}"
    base = url.rstrip("/")
    pages = [
        url,
        base + "/contact",
        base + "/contactpagina",
        base + "/contact-us",
        base + "/over-ons",
        base + "/over-de-club",
        base + "/about",
        base + "/team",
        base + "/ledenadministratie",
        base + "/secretariaat",
        base + "/bestuur",
    ]
    email_pattern = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
    junk = ["example", "sentry", "wix", "wordpress", "schema", ".png", ".jpg", "noreply",
            "privacy", "squarespace", "shopify", "weebly", "jimdo", "hostinger",
            "google", "facebook", "instagram", "twitter", "support@", "no-reply", "donotreply"]
    fake_contacts = {"to let", "te huur", "te koop", "for sale", "for rent", "n/a", "info",
                     "contact", "admin", "unknown", "webmaster", "hello", "team", "service"}

    name_patterns = [
        re.compile(r"(?:eigenaar|owner|directeur|manager|contact|ceo|founder)[:\s]+([A-Z][a-z]+(?:\s[A-Z][a-z]+)?)", re.IGNORECASE),
        re.compile(r"(?:Mijn naam is|I am|I'm)\s+([A-Z][a-z]+(?:\s[A-Z][a-z]+)?)", re.IGNORECASE),
    ]
    email, contact = "", ""

    # First load the homepage to discover additional contact links
    try:
        home_r = requests.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
        home_soup = BeautifulSoup(home_r.text, "html.parser")
        final_base = home_r.url.rstrip("/").rsplit("/", 1)[0] if "/" in home_r.url else home_r.url.rstrip("/")
        for a in home_soup.find_all("a", href=True):
            href = a["href"].lower()
            if any(kw in href for kw in ["contact", "over-ons", "about", "bestuur", "secretariaat"]):
                full = a["href"] if a["href"].startswith("http") else final_base + "/" + a["href"].lstrip("/")
                if full not in pages:
                    pages.append(full)
    except Exception:
        pass

    visited = set()
    for page_url in pages:
        if page_url in visited:
            continue
        visited.add(page_url)
        try:
            r = requests.get(page_url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if r.status_code >= 400:
                continue
            soup_email = BeautifulSoup(r.text, "html.parser")
            found_emails = set(e for e in email_pattern.findall(r.text) if not any(j in e.lower() for j in junk))
            for a in soup_email.find_all("a", href=True):
                href = a["href"]
                if href.startswith("mailto:"):
                    candidate = href[7:].split("?")[0].strip()
                    if candidate and "@" in candidate and not any(j in candidate.lower() for j in junk):
                        found_emails.add(candidate)
            if found_emails and not email:
                email = next(iter(found_emails))
            if not contact:
                text = soup_email.get_text()
                for pat in name_patterns:
                    m = pat.search(text)
                    if m:
                        candidate = m.group(1).strip()
                        words = candidate.split()
                        if (1 <= len(words) <= 3
                                and all(len(w) <= 20 for w in words)
                                and not any(c.isdigit() for c in candidate)
                                and candidate.lower() not in fake_contacts):
                            contact = candidate
                        break
            if email and contact:
                break
        except Exception:
            continue

    # Playwright fallback — voor JS-rendered sites zoals KNLTB.Club
    if not email and _acquire_browser_slot(f"email:{url}"):
        try:
            from playwright.sync_api import sync_playwright
            playwright_pages = [url, base + "/bestuur", base + "/contact", base + "/commissies", base + "/over-ons"]
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True, args=_CHROMIUM_ARGS)
                try:
                    context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
                    context.set_default_timeout(10000)
                    pw_page = context.new_page()
                    for page_url in playwright_pages:
                        if email:
                            break
                        try:
                            pw_page.goto(page_url, wait_until="domcontentloaded", timeout=15000)
                            pw_page.wait_for_timeout(1000)
                            html = pw_page.content()
                            found_emails = set(e for e in email_pattern.findall(html) if not any(j in e.lower() for j in junk))
                            # Ook mailto links ophalen via Playwright
                            for href in pw_page.eval_on_selector_all("a[href^='mailto:']", "els => els.map(e => e.href)"):
                                candidate = href.replace("mailto:", "").split("?")[0].strip()
                                if candidate and "@" in candidate and not any(j in candidate.lower() for j in junk):
                                    found_emails.add(candidate)
                            if found_emails:
                                email = next(iter(found_emails))
                        except Exception:
                            continue
                finally:
                    try:
                        browser.close()
                    except Exception:
                        pass
        except Exception as e:
            log.warning("Playwright e-mail-fallback mislukt voor %s: %s", url, e)
        finally:
            _release_browser_slot()

    return email, contact


# ── Board scraper for sports clubs ───────────────────────────────────────────

def find_board_page(website: str) -> str:
    """Find the URL of the board/about page on a sports club website."""
    if not website:
        return ""

    url = website if website.startswith("http") else f"https://{website}"
    base = url.rstrip("/")

    BOARD_KEYWORDS = [
        "hoofdbestuur", "bestuur", "over-ons", "over ons", "organisatie",
        "vereniging", "het-bestuur", "commissie", "about", "de-club",
        "ons team", "ons bestuur", "wie zijn wij", "wie zijn we",
        "mensen achter", "de mensen", "team", "vrijwilligers",
        "kader", "staf", "technische staf", "management",
        "contact personen", "contactpersonen", "de vereniging",
        "club info", "clubinfo", "over de club", "onze club",
        "lidmaatschap", "structuur", "leiding", "directie",
    ]

    candidate_paths = [
        "/organisatie/hoofdbestuur", "/bestuur", "/over-ons", "/organisatie",
        "/het-bestuur", "/vereniging", "/de-club", "/club/bestuur", "/about",
        "/team", "/ons-team", "/wie-zijn-wij", "/contact-personen",
        "/over-de-club", "/de-vereniging", "/onze-club", "/clubinfo",
        "/club/organisatie", "/vereniging/bestuur", "/about/team",
    ]

    try:
        r = requests.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code < 400:
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                link_text = (a.get_text(" ", strip=True) + " " + href).lower()
                if any(kw in link_text for kw in BOARD_KEYWORDS):
                    if href.startswith("http"):
                        return href
                    elif href.startswith("/"):
                        return base + href
                    elif not href.startswith("#") and not href.startswith("mailto"):
                        return base + "/" + href.lstrip("/")
    except Exception:
        pass

    # Fallback: try common paths
    for path in candidate_paths:
        candidate = base + path
        try:
            r = requests.head(candidate, timeout=5, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if r.status_code < 400:
                return candidate
        except Exception:
            continue

    return ""


# ── Events scraper ────────────────────────────────────────────────────────────

def scrape_events(city: str, max_results: int) -> list:
    events = []
    seen = set()

    sources = [
        f"https://www.eventbrite.nl/d/netherlands--{city.lower()}/events/",
        f"https://www.google.com/search?q=evenementen+{city}+2025+2026",
        f"https://www.uitagenda.nl/agenda/{city.lower()}",
    ]

    if not _acquire_browser_slot(f"events:{city}"):
        return events
    try:
        return _scrape_events_inner(city, max_results, events, seen, sources)
    finally:
        _release_browser_slot()


def _scrape_events_inner(city: str, max_results: int, events: list, seen: set, sources: list) -> list:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=_CHROMIUM_ARGS)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            viewport={"width": 1280, "height": 800},
        )
        context.set_default_timeout(10000)
        page = context.new_page()

        # Try Eventbrite
        try:
            page.goto(sources[0], wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)

            for text in ["Alles accepteren", "Accept", "I agree"]:
                try:
                    page.click(f'button:has-text("{text}")', timeout=2000)
                    page.wait_for_timeout(500)
                    break
                except Exception:
                    pass

            cards = page.query_selector_all('article, [data-testid="event-card"], .eds-event-card')
            for card in cards[:max_results]:
                try:
                    title_el = card.query_selector('h2, h3, .eds-event-card__formatted-name')
                    date_el = card.query_selector('time, .eds-event-card__sub-title, [data-testid*="date"]')
                    loc_el = card.query_selector('[data-testid*="location"], .card-text--truncated__one')
                    link_el = card.query_selector('a')

                    title = title_el.inner_text().strip() if title_el else ""
                    date = date_el.inner_text().strip() if date_el else ""
                    location = loc_el.inner_text().strip() if loc_el else city
                    link = link_el.get_attribute("href") if link_el else ""

                    if title and title not in seen:
                        seen.add(title)
                        events.append({
                            "id": str(uuid.uuid4()),
                            "title": title,
                            "date": date,
                            "location": location,
                            "city": city,
                            "source": "Eventbrite",
                            "link": link,
                            "organizer": "",
                            "email": "",
                            "phone": "",
                            "found_at": datetime.now().isoformat(),
                        })
                except Exception:
                    continue
        except Exception:
            pass

        # Try Google search for local events
        if len(events) < max_results:
            try:
                page.goto(
                    f"https://www.google.com/search?q=evenementen+{city}+2025+2026&hl=nl",
                    wait_until="domcontentloaded", timeout=30000
                )
                page.wait_for_timeout(2000)

                # Google events snippet
                event_cards = page.query_selector_all('[data-attrid*="event"], .YOGjf, .lu_map_section')
                for card in event_cards[:max_results - len(events)]:
                    try:
                        title_el = card.query_selector('div[role="heading"], h3, .YOGjf')
                        date_el = card.query_selector('.cEZxRc, time, .OGA8xd')
                        title = title_el.inner_text().strip() if title_el else ""
                        date = date_el.inner_text().strip() if date_el else ""
                        if title and title not in seen:
                            seen.add(title)
                            events.append({
                                "id": str(uuid.uuid4()),
                                "title": title,
                                "date": date,
                                "location": city,
                                "city": city,
                                "source": "Google",
                                "link": "",
                                "organizer": "",
                                "email": "",
                                "phone": "",
                                "found_at": datetime.now().isoformat(),
                            })
                    except Exception:
                        continue
            except Exception:
                pass

        try:
            browser.close()
        except Exception:
            pass

    return events


# ── Cold email generator ──────────────────────────────────────────────────────

def _logo_display_size(logo_url: str) -> tuple:
    """Bepaal de weergavemaat van het logo in de e-mail: max 70 hoog, max 160 breed,
    met behoud van verhouding — zodat brede logo's niet in een vierkant geperst worden."""
    try:
        import io as _io

        from PIL import Image

        r = requests.get(logo_url, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
        img = Image.open(_io.BytesIO(r.content))
        w, h = img.size
        if w > 0 and h > 0:
            scale = min(70 / h, 160 / w)
            return max(int(w * scale), 1), max(int(h * scale), 1)
    except Exception:
        pass
    return 70, 70

SPORT_KEYWORDS_MAP = {
    "voetbal":       ["voetbal", "voetbalclub", "voetbalvereniging"],
    "hockey":        ["hockey", "hockeyclub", "hockeyvereniging"],
    "tennis":        ["tennis", "padel", "padelclub", "tennisclub", "tennisvereniging"],
    "golf":          ["golf", "golfclub", "golfvereniging"],
}


def _apply_email_template(raw_html: str, name: str, logo_url: str, is_sport: bool) -> str:
    import html as html_mod
    import re as _re
    clean_name = _re.sub(r'[^\x00-\x7FÀ-ɏḀ-ỿ]', '', name).strip()
    escaped_name = html_mod.escape(clean_name)

    # Replace name placeholder (sport or eten)
    raw_html = raw_html.replace("&lt;naam vereniging&gt;", escaped_name)
    raw_html = raw_html.replace("<naam vereniging>", escaped_name)
    raw_html = raw_html.replace("[NAAM ZAAK]", escaped_name)

    soup = BeautifulSoup(raw_html, "html.parser")

    # Restore green background on header and footer rows (mammoth strips table bg colors)
    heineken_imgs = soup.find_all("img", alt="Heineken Horeca")
    for img in heineken_imgs:
        # Walk up: img → td → tr → table → tr (this is the section row we want to color)
        inner_tr = img.find_parent("tr")
        if not inner_tr:
            continue
        inner_table = inner_tr.find_parent("table")
        if not inner_table:
            continue
        section_tr = inner_table.find_parent("tr")
        if not section_tr:
            continue
        section_td = section_tr.find_parent("td")
        outer_table = section_td.find_parent("table") if section_td else None
        outer_tr = outer_table.find_parent("tr") if outer_table else None
        # Color the section td that wraps the logo
        if section_td:
            section_td["style"] = "background-color:#005826;padding:0"
        # Also color the outer tr if it exists
        if outer_tr:
            outer_tr["style"] = "background-color:#005826"
        # Make the nested table fill the width
        if inner_table:
            inner_table["style"] = "width:100%"
            section_tr_parent = inner_table
            # color all tds in the header/footer row
            for td in inner_tr.find_all("td"):
                td["style"] = "background-color:#005826;padding:8px 16px"

    # Insert club logo to the right of the email title
    if is_sport and logo_url:
        title_p = None
        # Zoek eerst in <p> tags direct
        for p in soup.find_all("p"):
            if "kennismaking" in p.get_text().lower():
                title_p = p
                break
        # Fallback: zoek de diepste <p> die via parent een td/div bevat met de tekst
        if not title_p:
            for tag in soup.find_all(["td", "div"]):
                if "kennismaking" in tag.get_text().lower():
                    inner_p = tag.find("p")
                    title_p = inner_p if inner_p else tag
                    break
        if title_p:
            # Wrap title + logo in a flex row so they sit side by side
            wrapper = soup.new_tag("div")
            wrapper["style"] = "display:flex;align-items:center;justify-content:space-between;gap:24px;width:100%"
            # Move the title <p> content into a span inside the wrapper
            title_span = soup.new_tag("span")
            title_span["style"] = "flex:1"
            for child in list(title_p.children):
                title_span.append(child.extract())
            logo_w, logo_h = _logo_display_size(logo_url)
            club_logo_tag = soup.new_tag("img")
            club_logo_tag["src"] = logo_url
            club_logo_tag["alt"] = f"{escaped_name} logo"
            club_logo_tag["width"] = str(logo_w)
            club_logo_tag["height"] = str(logo_h)
            club_logo_tag["style"] = (
                f"width:{logo_w}px;height:{logo_h}px;object-fit:contain;flex-shrink:0;"
                "background:white;border-radius:6px;padding:4px;border:1px solid #ddd;margin-left:40px"
            )
            wrapper.append(title_span)
            wrapper.append(club_logo_tag)
            title_p.append(wrapper)

    # Set explicit widths on all images so Outlook doesn't show them at full resolution
    for img in soup.find_all("img"):
        alt = img.get("alt", "")
        src = img.get("src", "")
        # Skip club logo (already sized) and tiny icons
        if img.get("width") and int(str(img.get("width", "0")).replace("px","") or 0) <= 80:
            continue
        if "logo_url" in src or alt == f"{escaped_name} logo":
            continue
        # Heineken header/footer logos: keep natural width but cap height
        if alt == "Heineken Horeca":
            img["style"] = img.get("style", "") + ";max-width:200px;height:auto"
            img["width"] = "200"
        else:
            # Content images (sport cards, YouTube thumbnail): fit email width
            img["style"] = img.get("style", "") + ";max-width:560px;width:100%;height:auto"
            img["width"] = "560"

    return str(soup)


# ── Menukaart → bier-spijscombinatie (eetgelegenheden) ───────────────────────

# Volgorde = prioriteit bij gelijke score. Trefwoorden matchen ook samenstellingen
# ("vis" matcht "visgerechten") maar niet middenin een woord ("advies" niet).
_BEER_PAIRINGS = [
    ("Pizza", ["pizza"], "Birra Moretti",
     "Birra Moretti is dé Italiaanse klassieker: zacht mout en subtiele hop, precies wat een pizza uit de oven vraagt."),
    ("Pasta", ["pasta", "spaghetti", "lasagne", "risotto", "gnocchi"], "Birra Moretti",
     "De zachte, licht moutige smaak van Birra Moretti sluit van nature aan bij Italiaanse gerechten."),
    ("Spareribs", ["spareribs", "sparerib", "ribs", "pulled pork", "bbq", "barbecue"], "Texels Skuumkoppe",
     "Het donkere tarwebier Texels Skuumkoppe heeft karamelzoete tonen die de rooksmaak van gegrild vlees versterken."),
    ("Burgers", ["burger", "hamburger", "cheeseburger"], "Lagunitas IPA",
     "De frisse hopbitterheid van Lagunitas IPA snijdt mooi door de volle smaak van een burger heen."),
    ("Steak", ["steak", "biefstuk", "entrecote", "tournedos", "ribeye", "bavette"], "Affligem Dubbel",
     "De rijke karameltonen van Affligem Dubbel geven gegrild rundvlees extra diepgang."),
    ("Saté", ["saté", "satay"], "Oedipus Mannenliefde",
     "Oedipus Mannenliefde, gebrouwen met citroengras en Szechuanpeper, geeft een frisse tegenhanger aan de rijke pindasaus."),
    ("Aziatische gerechten", ["curry", "thais", "thaise", "ramen", "noodles", "wok", "sushi", "poke", "dim sum"], "Oedipus Mannenliefde",
     "Het kruidige, frisse Oedipus Mannenliefde is gebrouwen met citroengras en past daardoor verrassend goed bij Aziatische smaken."),
    ("Visgerechten", ["zalm", "kabeljauw", "tonijn", "garnalen", "mosselen", "kibbeling", "vis", "zeebaars", "gamba"], "Affligem Blond",
     "Het licht fruitige Affligem Blond begeleidt vis zonder de fijne smaak te overheersen."),
    ("Stoofgerechten", ["stoofvlees", "stoofpot", "stoverij", "sudderlap"], "Affligem Dubbel",
     "Affligem Dubbel heeft dezelfde donkere, zoete diepte als een goed stoofgerecht — ze versterken elkaar."),
    ("Kaasplank", ["kaasplank", "kaasfondue", "borrelplank"], "Affligem Tripel",
     "De kruidige volheid van Affligem Tripel is een klassieke combinatie met gerijpte kazen."),
    ("Desserts", ["dessert", "chocolade", "appeltaart", "tiramisu", "cheesecake"], "Mort Subite Kriek",
     "Het friszoete kersenbier Mort Subite Kriek komt naast een dessert perfect tot zijn recht."),
]

_MENU_LINK_WORDS = ["menu", "menukaart", "kaart", "lunch", "diner", "dinerkaart", "gerechten"]


def find_menu_pairing(website: str) -> tuple | None:
    """Zoek op de website (en menupagina's) het meest genoemde gerecht en koppel
    er een Heineken-speciaalbier aan. Geeft (gerecht, bier, uitleg) of None."""
    if not website:
        return None
    url = website if website.startswith("http") else f"https://{website}"
    texts = []
    menu_pages = []
    try:
        r = requests.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
        if r.status_code < 400:
            soup = BeautifulSoup(r.text, "html.parser")
            texts.append(soup.get_text(" ", strip=True))
            for a in soup.find_all("a", href=True):
                label = f"{a.get_text()} {a['href']}".lower()
                if any(w in label for w in _MENU_LINK_WORDS):
                    full = urljoin(r.url, a["href"])
                    if full not in menu_pages and not full.lower().endswith(".pdf"):
                        menu_pages.append(full)
    except Exception:
        return None
    # Max 2 menupagina's ophalen: elke fetch kost tijd binnen het jobbudget
    for page_url in menu_pages[:2]:
        try:
            r = requests.get(page_url, timeout=6, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if r.status_code < 400 and "text/html" in r.headers.get("content-type", ""):
                texts.append(BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True))
        except Exception:
            continue

    full_text = " ".join(texts).lower()
    if not full_text.strip():
        return None
    best = None
    best_hits = 0
    for dish, keywords, beer, why in _BEER_PAIRINGS:
        hits = sum(len(re.findall(rf"\b{re.escape(kw)}", full_text)) for kw in keywords)
        if hits > best_hits:
            best, best_hits = (dish, beer, why), hits
    return best


# Letterlijke blokken uit het eten-sjabloon (static/email-templates/eten.html)
_PAIRING_BLOCK_2 = "<p><strong>[Gerecht 2] - [Speciaalbier 2]</strong></p><p>Korte toelichting.</p>"
_PAIRING_SECTION = (
    "<p><strong>Welke bieren passen goed bij jullie gerechten</strong></p>"
    "<table><tr><td><p><strong>[Gerecht 1] - [Speciaalbier 1]</strong></p>"
    "<p>Korte toelichting.</p>" + _PAIRING_BLOCK_2 + "</td></tr></table>"
)


def _fill_menu_pairing(raw_html: str, website: str) -> str:
    """Vul het bier-spijsblok van de eetmail met een gerecht van de menukaart.

    Geen gerecht gevonden → hele blok verwijderen, zodat er nooit een lege
    sjabloontekst als "[Gerecht 1]" in een verstuurde mail staat."""
    try:
        pairing = find_menu_pairing(website)
    except Exception:
        log.exception("Menukoppeling mislukt voor %s", website)
        pairing = None
    if not pairing:
        return raw_html.replace(_PAIRING_SECTION, "")
    dish, beer, why = pairing
    raw_html = raw_html.replace(_PAIRING_BLOCK_2, "")
    raw_html = raw_html.replace("[Gerecht 1]", dish).replace("[Speciaalbier 1]", beer)
    return raw_html.replace("Korte toelichting.", why, 1)


def generate_cold_email(business: dict) -> str:
    name = business["name"]
    niche = business.get("niche", "").lower()
    lead_type = business.get("type", "business")
    logo_url = business.get("logo_url", "")

    if lead_type == "sport":
        sport = "overig"
        for key, keywords in SPORT_KEYWORDS_MAP.items():
            if any(kw in niche for kw in keywords):
                sport = key
                break
        template_key = sport if sport in EMAIL_TEMPLATES else "voetbal"
        raw_html = EMAIL_TEMPLATES.get(template_key, "")
        if raw_html:
            return _apply_email_template(raw_html, name, logo_url, is_sport=True)
    else:
        raw_html = EMAIL_TEMPLATES.get("eten", "")
        if raw_html:
            raw_html = _fill_menu_pairing(raw_html, business.get("website", ""))
            return _apply_email_template(raw_html, name, logo_url, is_sport=False)

    return f"<p>Email template niet beschikbaar voor {name}</p>"


def find_website_by_name(name: str) -> str:
    """Zoek de officiële website van een club via DuckDuckGo."""
    import time
    from ddgs import DDGS
    skip_domains = ["facebook.com", "instagram.com", "twitter.com", "linkedin.com",
                    "youtube.com", "wikipedia.org", "yelp.com", "tripadvisor",
                    "booking.com", "toernooiinfo", "sportlink", "clubwebsite",
                    "knltb.nl", "nevobo.nl", "knhb.nl", "knvb.nl", "google.",
                    "thuisbezorgd", "iens.nl", "spoj.com", "wanderlog"]
    sport_words = ["tennis", "sport", "vereniging", "club", "padel", "hockey",
                   "voetbal", "golf", "zwem", "atletiek", "tv.", "tc.", ".tv"]
    queries = [
        f'"{name}" site:*.nl',
        f"{name} officiële website tennis",
    ]
    try:
        with DDGS() as ddgs:
            for query_str in queries:
                try:
                    results = list(ddgs.text(query_str, max_results=10))
                    time.sleep(1.5)
                    for r in results:
                        href = r.get("href") or r.get("url", "")
                        title = (r.get("title", "") + " " + r.get("body", "")).lower()
                        if not href.startswith("http"):
                            continue
                        if any(d in href for d in skip_domains):
                            continue
                        href_low = href.lower()
                        if any(w in href_low or w in title for w in sport_words):
                            return href
                except Exception:
                    time.sleep(2)
                    continue
    except Exception:
        pass
    return ""


def guess_club_website(name: str) -> str:
    """Probeer voorspelbare URL-patronen voor Nederlandse clubs en horeca. Verifieert dat de site relevant is."""
    bare = _CLUB_PREFIXES.sub("", name).strip().lower()
    bare = re.sub(r"^(de|het|den|'t)\s+", "", bare, flags=re.IGNORECASE).strip()
    bare = re.sub(r"[^a-z0-9]", "", bare)
    if not bare or len(bare) < 3:
        return ""
    verify_words = [
        "tennis", "sport", "club", "vereniging", "padel", "baan", "leden", "knltb",
        "voetbal", "hockey", "golf", "zwemmen", "atletiek",
        "restaurant", "cafe", "café", "horeca", "menu", "reserveren", "eten",
    ]
    candidates = [
        f"https://www.{bare}.nl",
        f"https://www.tv{bare}.nl",
        f"https://www.tc{bare}.nl",
        f"https://www.sv{bare}.nl",
        f"https://www.vv{bare}.nl",
        f"https://www.hv{bare}.nl",
        f"https://www.tennis{bare}.nl",
        f"https://www.{bare}tennis.nl",
        f"https://www.ltv{bare}.nl",
        f"https://www.{bare}tc.nl",
        f"https://www.{bare}tv.nl",
        f"https://www.cafe{bare}.nl",
        f"https://www.restaurant{bare}.nl",
    ]
    for url in candidates:
        try:
            r = requests.get(url, timeout=5, allow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code < 400:
                text = r.text.lower()
                if any(w in text for w in verify_words):
                    return url
        except Exception:
            continue
    return ""


# ── Background jobs ───────────────────────────────────────────────────────────

def _scrape_maps_with_watchdog(niche: str, city: str, count: int, timeout_s: int = 240) -> list:
    """Draai de Maps-scrape met een harde timeout, zodat een vastgelopen browser nooit de hele job blokkeert."""
    from concurrent.futures import TimeoutError as _FuturesTimeout
    worker_info = {}
    # Gedeelde lijst: de scraper vult deze incrementeel, zodat we bij een
    # ingreep van de watchdog de al gevonden plaatsen kunnen redden
    partial_results: list = []

    def _run():
        worker_info["tid"] = threading.get_ident()
        return scrape_google_maps(niche, city, count, results=partial_results)

    ex = ThreadPoolExecutor(max_workers=1)
    fut = ex.submit(_run)
    try:
        return fut.result(timeout=timeout_s)
    except _FuturesTimeout:
        log.error("Maps-scrape watchdog: vastgelopen na %ds voor %r / %r", timeout_s, niche, city)
        # Houdt de vastgelopen scrape het browser-slot vast? Dan is er maar
        # één browser actief (slots=1) en kunnen we die veilig hard opruimen —
        # dat laat de hangende playwright-call crashen zodat het slot vrijkomt.
        with _slot_lock:
            holds_slot = worker_info.get("tid") in _slot_holders
        if holds_slot:
            pids = _list_chromium_pids()
            log.warning("Watchdog: vastgelopen scrape houdt het browser-slot vast — %d chromium-processen opruimen", len(pids))
            for pid in pids:
                try:
                    os.kill(pid, 9)
                except Exception:
                    pass
        # De kill laat de hangende playwright-call crashen; geef de worker
        # even de tijd om netjes terug te keren met wat al binnen is
        try:
            return fut.result(timeout=15)
        except Exception:
            pass
        if partial_results:
            log.warning("Watchdog: %d deelresultaten gered voor %r / %r", len(partial_results), niche, city)
            return list(partial_results)
        raise RuntimeError("De zoek-browser op de server reageert niet — probeer het over een paar minuten opnieuw")
    finally:
        ex.shutdown(wait=False, cancel_futures=True)

def _names_similar(a: str, b: str, threshold: float = 0.82) -> bool:
    """True als twee namen waarschijnlijk dezelfde zaak/club zijn."""
    a, b = a.lower().strip(), b.lower().strip()
    # Strip suffixen VOOR vergelijking
    for suffix in [" b.v.", " bv", " v.o.f.", " vof", " stichting", " vereniging", " e.o.", " e.v."]:
        a = a.removesuffix(suffix).strip()
        b = b.removesuffix(suffix).strip()
    if a == b:
        return True
    return SequenceMatcher(None, a, b).ratio() >= threshold


def scrape_lead_details(b: dict, is_sport: bool = False):
    """Scrape alle details van een lead: email, logo, bestuurspagina, screenshot, cold email."""
    website = b.get("website", "")
    b["website_status"] = check_website(website)
    b["screenshot"] = take_screenshot(website, b["id"])
    b["email"], b["contact_person"] = find_email_and_contact(website)
    b["logo_url"] = find_logo(website)
    if not b.get("phone"):
        b["phone"] = find_phone_from_website(website)
    if not b.get("address"):
        b["address"] = find_address_from_website(website)
    if is_sport:
        b["board_page"] = find_board_page(website)
    email_html = generate_cold_email(b)
    save_email(b["id"], email_html)
    b["cold_email"] = ""


def run_search_job(job_id: str, niche: str, city: str, max_results: int, force_type: str = None, radius: int = 0, user_id: int = None):
    jobs[job_id] = {"status": "running", "progress": 0, "message": "Google Maps doorzoeken..."}
    log.info("Zoekjob %s gestart: niche=%r stad=%r max=%d", job_id, niche, city, max_results)
    _wait_until_slot_free(job_id)
    # Wachttijd op andere gebruikers telt niet mee voor het tijdsbudget van de job
    start_time = time.time()
    # Ruim genoeg voor serieel browserwerk (1 browser tegelijk op de server)
    MAX_JOB_SECONDS = 300
    try:
        if force_type == "sport":
            sport = True
        elif force_type == "business":
            sport = False
        else:
            sport = is_sport(niche)

        existing_leads = load_leads(user_id)
        existing_names_by_city: dict[str, list[str]] = {}
        for l in existing_leads:
            c = l.get("city", "").lower().strip()
            existing_names_by_city.setdefault(c, []).append(l["name"])

        # ×2 (was ×3): elke kandidaat kost serieel browsertijd, dus niet te veel ophalen
        fetch_count = max_results * 2
        search_city = city or _extract_city_from_name(niche) or ""
        jobs[job_id].update({"progress": 5, "message": f"Google Maps doorzoeken in {search_city}..."})
        businesses = _scrape_maps_with_watchdog(niche, search_city, fetch_count)
        seen_names = {b["name"].lower() for b in businesses}

        # Bij radius: zoek in omliggende plaatsen
        if radius > 0:
            nearby = get_nearby_places(search_city, radius)
            for i, place in enumerate(nearby):
                if len(businesses) >= fetch_count:
                    break
                pct = 10 + int((i / max(len(nearby), 1)) * 20)
                jobs[job_id].update({"progress": pct, "message": f"Zoeken in {place} ({radius}km omgeving)..."})
                extra = _scrape_maps_with_watchdog(niche, place, max(5, fetch_count // len(nearby) if nearby else fetch_count))
                for b in extra:
                    if b["name"].lower() not in seen_names:
                        businesses.append(b)
                        seen_names.add(b["name"].lower())
        # Als weinig resultaten zonder radius: zoek ook in de gemeente
        elif len(businesses) < max_results:
            gemeente = get_gemeente(search_city)
            if gemeente:
                jobs[job_id].update({"progress": 15, "message": f"Weinig resultaten in {city}, ook zoeken in gemeente {gemeente}..."})
                extra = _scrape_maps_with_watchdog(niche, gemeente, fetch_count)
                for b in extra:
                    if b["name"].lower() not in seen_names:
                        businesses.append(b)
                        seen_names.add(b["name"].lower())

        jobs[job_id].update({"progress": 30, "message": "Gevonden clubs verwerken..."})

        if force_type:
            for b in businesses:
                b["type"] = force_type

        def _is_duplicate(b: dict) -> bool:
            b_city = b.get("city", city).lower().strip()
            for existing_name in existing_names_by_city.get(b_city, []):
                if _names_similar(b["name"], existing_name):
                    return True
            return False

        new_businesses = [b for b in businesses if not _is_duplicate(b)]
        # Haal meer op dan nodig zodat we leads zonder website kunnen overslaan
        candidate_pool = new_businesses[:max_results * 2]
        total_candidates = len(candidate_pool)

        if total_candidates == 0:
            jobs[job_id] = {"status": "done", "progress": 100, "message": "Geen nieuwe leads gevonden — alles al bekend. Probeer een andere zoekterm of stad.", "count": 0}
            return

        completed = [0]
        saved_leads = []
        lock = threading.Lock()

        def process_lead(b):
            # Genoeg leads binnen? Dan niets meer scrapen — scheelt browsers en geheugen
            with lock:
                if len(saved_leads) >= max_results:
                    return b
            scrape_lead_details(b, is_sport=sport)
            with lock:
                completed[0] += 1
                # Sla leads zonder website over
                if not b.get("website"):
                    pct = 30 + int((len(saved_leads) / max(max_results, 1)) * 70)
                    jobs[job_id]["progress"] = pct
                    jobs[job_id]["message"] = f"{len(saved_leads)}/{max_results} leads gevonden..."
                    return b
                if len(saved_leads) < max_results and add_lead(b, user_id):
                    saved_leads.append(b)
                pct = 30 + int((len(saved_leads) / max(max_results, 1)) * 70)
                jobs[job_id]["progress"] = pct
                jobs[job_id]["message"] = f"{len(saved_leads)}/{max_results} leads gevonden..."
            return b

        # Max 2 tegelijk: elke lead start meerdere Chromium-instanties en de server heeft beperkt geheugen
        executor = ThreadPoolExecutor(max_workers=2)
        try:
            futures = {executor.submit(process_lead, b): b for b in candidate_pool}
            for future in as_completed(futures, timeout=MAX_JOB_SECONDS):
                if len(saved_leads) >= max_results:
                    break
                if time.time() - start_time > MAX_JOB_SECONDS:
                    jobs[job_id] = {"status": "done", "progress": 100, "message": f"Timeout — {len(saved_leads)} leads gevonden", "count": len(saved_leads)}
                    return
                try:
                    future.result()
                except Exception:
                    pass
        finally:
            # Niet wachten op de rest: nog niet gestarte kandidaten annuleren,
            # anders blijft de job op 100% "hangen" tot alle 30 klaar zijn
            executor.shutdown(wait=False, cancel_futures=True)

        actual_saved = len(saved_leads)
        skipped = len(businesses) - actual_saved
        skip_msg = f" ({skipped} al bekend overgeslagen)" if skipped > 0 else ""
        jobs[job_id] = {"status": "done", "progress": 100, "message": f"{actual_saved} nieuwe leads opgeslagen!{skip_msg}", "count": actual_saved}
        log.info("Zoekjob %s klaar: %d leads opgeslagen", job_id, actual_saved)
    except Exception as e:
        log.exception("Zoekjob %s mislukt", job_id)
        jobs[job_id] = {"status": "error", "progress": 0, "message": str(e)}


def run_events_job(job_id: str, city: str, max_results: int):
    jobs[job_id] = {"status": "running", "progress": 0, "message": f"Evenementen zoeken in {city}..."}
    _wait_until_slot_free(job_id)
    try:
        events = scrape_events(city, max_results)
        all_events = load_events()
        existing_event_keys = {e.get("title", e.get("name", "")).lower().strip() for e in all_events}
        new_events = [e for e in events if e["title"].lower().strip() not in existing_event_keys]
        all_events = new_events + all_events
        save_events(all_events)
        jobs[job_id] = {"status": "done", "progress": 100, "message": f"{len(new_events)} evenementen gevonden!", "count": len(new_events)}
    except Exception as e:
        log.exception("Events-job mislukt voor %s", city)
        jobs[job_id] = {"status": "error", "progress": 0, "message": str(e)}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('index'))
        return render_template("login.html", error="Ongeldig e-mailadres of wachtwoord")
    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register_page():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if User.query.filter_by(email=email).first():
            return render_template("register.html", error="Dit e-mailadres is al in gebruik")
        user = User(name=name, email=email)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for('index'))
    return render_template("register.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for('login_page'))


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/api/search", methods=["POST"])
@login_required
def search():
    data = request.json
    niche = data.get("niche", "").strip()
    city = data.get("city", "").strip()
    # Max 20: de job wordt na 180s afgekapt, meer dan ~20 leads passen daar toch niet in
    max_results = min(int(data.get("max_results", data.get("max", 10))), 20)
    force_type = data.get("force_type")
    radius = int(data.get("radius", 0))
    if not niche:
        return jsonify({"error": "Vul een naam of niche in"}), 400
    job_id = str(uuid.uuid4())
    _job_executor.submit(run_search_job, job_id, niche, city, max_results, force_type, radius, current_user.id)
    return jsonify({"job_id": job_id})


@app.route("/api/events/search", methods=["POST"])
@login_required
def events_search():
    data = request.json
    city = data.get("city", "").strip()
    max_results = min(int(data.get("max", 10)), 50)
    if not city:
        return jsonify({"error": "Vul een stad in"}), 400
    job_id = str(uuid.uuid4())
    _job_executor.submit(run_events_job, job_id, city, max_results)
    return jsonify({"job_id": job_id})


@app.route("/api/job/<job_id>")
def job_status(job_id):
    return jsonify(jobs.get(job_id, {"status": "not_found"}))


@app.route("/api/debug")
@login_required
def debug_info():
    """Serverstatus: geheugen, browserprocessen en recente logs — voor probleemdiagnose."""
    now = time.time()
    with _slot_lock:
        slot_info = [{"taak": what, "bezig_sinds_s": int(now - started)} for what, started in _slot_holders.values()]
    info = {
        "threads": len(threading.enumerate()),
        "browser_slots": {"totaal": _BROWSER_SLOTS_TOTAL, "bezet": slot_info},
        "jobs_recent": {k: {"status": v.get("status"), "progress": v.get("progress"), "message": v.get("message")}
                        for k, v in list(jobs.items())[-10:]},
    }
    try:
        mem = {}
        with open("/proc/meminfo") as f:
            for line in f:
                if line.startswith(("MemTotal", "MemAvailable", "SwapFree")):
                    key, val = line.split(":", 1)
                    mem[key] = val.strip()
        info["meminfo"] = mem
    except Exception:
        pass
    for path, key in [("/sys/fs/cgroup/memory.current", "cgroup_memory_current"),
                      ("/sys/fs/cgroup/memory.max", "cgroup_memory_max"),
                      ("/sys/fs/cgroup/memory/memory.usage_in_bytes", "cgroup_memory_current"),
                      ("/sys/fs/cgroup/memory/memory.limit_in_bytes", "cgroup_memory_max")]:
        try:
            with open(path) as f:
                info[key] = f.read().strip()
        except Exception:
            pass
    try:
        proc_names = []
        for pid in os.listdir("/proc"):
            if pid.isdigit():
                try:
                    with open(f"/proc/{pid}/comm") as f:
                        proc_names.append(f.read().strip())
                except Exception:
                    pass
        info["totaal_processen"] = len(proc_names)
        levend = len(_list_chromium_pids())
        totaal_chromium = len(_list_chromium_pids(include_zombies=True))
        info["chromium_processen"] = levend
        info["chromium_zombies"] = totaal_chromium - levend
        info["node_processen"] = sum(1 for n in proc_names if n.lower().startswith("node"))
    except Exception:
        pass
    try:
        info["accounts"] = [
            {
                "naam": u.name,
                "email": u.email,
                "aangemaakt": u.created_at.strftime("%Y-%m-%d") if u.created_at else None,
                "aantal_leads": db.session.query(Lead.id).filter_by(user_id=u.id).count(),
            }
            for u in User.query.order_by(User.created_at).all()
        ]
    except Exception as e:
        info["accounts_fout"] = str(e)
    info["log"] = list(_log_buffer)
    return jsonify(info)


@app.route("/api/job/<job_id>/cancel", methods=["POST"])
@login_required
def cancel_job(job_id):
    if job_id in jobs and jobs[job_id].get("status") == "running":
        jobs[job_id] = {"status": "cancelled", "progress": 0, "message": "Gestopt door gebruiker", "count": 0}
    return jsonify({"ok": True})


@app.route("/api/leads")
@login_required
def get_leads():
    leads = load_leads(current_user.id)
    status_filter = request.args.get("status")
    type_filter = request.args.get("type")
    if status_filter:
        leads = [l for l in leads if l.get("website_status") == status_filter]
    if type_filter:
        leads = [l for l in leads if l.get("type") == type_filter]
    # Screenshots zitten in een aparte kolom en komen hier dus nooit mee —
    # de lijst blijft licht, ook op de server (dit veroorzaakte OOM-crashes)
    return jsonify(leads)


@app.route("/api/leads/<lead_id>/screenshot")
@login_required
def lead_screenshot(lead_id):
    import base64 as _b64
    row = Lead.query.filter_by(id=lead_id, user_id=current_user.id).first()
    shot = (row.screenshot or (row.data or {}).get("screenshot", "")) if row else ""
    if not shot:
        return "", 404
    if not shot.startswith("data:image"):
        # Oude leads verwijzen nog naar een los bestand
        return redirect(f"/screenshots/{shot}")
    header, b64data = shot.split(",", 1)
    mime = header.split(";")[0].replace("data:", "") or "image/png"
    resp = app.response_class(_b64.b64decode(b64data), mimetype=mime)
    resp.headers["Cache-Control"] = "private, max-age=3600"
    return resp


@app.route("/api/leads/<lead_id>", methods=["PATCH"])
@login_required
def update_lead(lead_id):
    leads = load_leads(current_user.id)
    data = request.json
    for lead in leads:
        if lead["id"] == lead_id:
            lead.update(data)
            break
    save_leads(leads, current_user.id)
    return jsonify({"ok": True})


@app.route("/api/leads/<lead_id>/status", methods=["PATCH"])
@login_required
def update_lead_status(lead_id):
    leads = load_leads(current_user.id)
    status = request.json.get("status", "nieuw")
    for lead in leads:
        if lead["id"] == lead_id:
            lead["status"] = status
            break
    save_leads(leads, current_user.id)
    return jsonify({"ok": True})


@app.route("/api/leads/<lead_id>/email")
@login_required
def get_lead_email(lead_id):
    html = load_email(lead_id)
    leads = load_leads(current_user.id)
    lead = next((l for l in leads if l["id"] == lead_id), None)
    if lead:
        logo_url = lead.get("logo_url", "")
        # Regenerate if email is missing, or has a logo available but email doesn't show it
        needs_regen = not html or (logo_url and 'logo_url' not in html and logo_url not in html)
        if needs_regen:
            html = generate_cold_email(lead)
            save_email(lead_id, html)
    return jsonify({"html": html})


@app.route("/api/leads/<lead_id>", methods=["DELETE"])
@login_required
def delete_lead(lead_id):
    leads = [l for l in load_leads(current_user.id) if l["id"] != lead_id]
    save_leads(leads, current_user.id)
    delete_email(lead_id)
    return jsonify({"ok": True})


@app.route("/api/leads/<lead_id>/rescrape", methods=["POST"])
@login_required
def rescrape_lead(lead_id):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "running", "progress": 0, "message": "Opnieuw scrapen..."}
    _wait_until_slot_free(job_id)

    uid = current_user.id

    def _run():
        try:
            leads = load_leads(uid)
            lead = next((l for l in leads if l["id"] == lead_id), None)
            if not lead:
                jobs[job_id] = {"status": "error", "message": "Lead niet gevonden"}
                return

            website = lead.get("website", "")

            # Als er nog geen website is, probeer die opnieuw te vinden
            if not website:
                jobs[job_id]["message"] = "Website opzoeken..."
                name = lead.get("niche") or lead.get("name", "")
                force_type = lead.get("type", "sport")
                extracted_city = _extract_city_from_name(name)
                bare_name = _CLUB_PREFIXES.sub("", name).strip()
                bare_name_no_de = re.sub(r"^(de|het|den|'t)\s+", "", bare_name, flags=re.IGNORECASE).strip()
                queries = [
                    (name, extracted_city),
                    (bare_name, extracted_city), (bare_name_no_de, extracted_city),
                    (name, ""), (bare_name, ""),
                ]
                first_result = None
                for q_name, q_city in queries:
                    if website or not q_name.strip():
                        break
                    try:
                        businesses = _scrape_maps_with_watchdog(q_name, q_city, 5, timeout_s=150)
                        match = next((b for b in businesses if _names_similar(name, b["name"], threshold=0.45)), None)
                        if match:
                            lead["name"] = match["name"]
                            website = match.get("website", "")
                            lead["phone"] = lead.get("phone") or match.get("phone", "")
                            lead["address"] = lead.get("address") or match.get("address", "")
                            lead["city"] = lead.get("city") or match.get("city", "")
                        elif businesses and not first_result:
                            first_result = businesses[0]
                    except Exception:
                        continue
                if not website and first_result:
                    website = first_result.get("website", "")
                    lead["phone"] = lead.get("phone") or first_result.get("phone", "")
                    lead["address"] = lead.get("address") or first_result.get("address", "")
                    lead["city"] = lead.get("city") or first_result.get("city", "")
                if not website:
                    website = guess_club_website(name)
                lead["website"] = website

            jobs[job_id]["progress"] = 30
            jobs[job_id]["message"] = "Screenshot maken..."
            if website:
                lead["screenshot"] = take_screenshot(website, lead_id)

            jobs[job_id]["progress"] = 50
            jobs[job_id]["message"] = "Email zoeken..."
            if website:
                email, contact = find_email_and_contact(website)
                if email:
                    lead["email"] = email
                if contact:
                    lead["contact_person"] = contact

            jobs[job_id]["progress"] = 70
            jobs[job_id]["message"] = "Telefoon en adres zoeken..."
            if website and not lead.get("phone"):
                lead["phone"] = find_phone_from_website(website)
            if website and not lead.get("address"):
                lead["address"] = find_address_from_website(website)

            jobs[job_id]["progress"] = 85
            jobs[job_id]["message"] = "Email template bijwerken..."
            if website and lead.get("type") == "sport" and not lead.get("board_page"):
                try:
                    lead["board_page"] = find_board_page(website)
                except Exception:
                    pass
            if website:
                # Altijd opnieuw zoeken: een eerder gekozen logo kan een wazig mini-icoontje zijn
                lead["logo_url"] = find_logo(website) or lead.get("logo_url", "")
            try:
                save_email(lead_id, generate_cold_email(lead))
            except Exception:
                pass

            with _leads_lock:
                all_leads = _load_leads_unsafe(uid)
                for i, l in enumerate(all_leads):
                    if l["id"] == lead_id:
                        all_leads[i] = lead
                        break
                _save_leads_unsafe(all_leads, uid)

            # has_screenshot bijwerken: een verse screenshot zit in lead["screenshot"]
            # maar gaat niet mee in het antwoord (te groot voor de job-JSON)
            lead["has_screenshot"] = bool(lead.get("screenshot") or lead.get("has_screenshot"))
            jobs[job_id] = {"status": "done", "progress": 100, "message": "Klaar!",
                            "lead": {k: v for k, v in lead.items() if k != "screenshot"}}
        except Exception as e:
            jobs[job_id] = {"status": "error", "message": str(e)}

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/events")
@login_required
def get_events():
    return jsonify(load_events())


@app.route("/api/events/<event_id>", methods=["DELETE"])
@login_required
def delete_event(event_id):
    events = [e for e in load_events() if e["id"] != event_id]
    save_events(events)
    return jsonify({"ok": True})


@app.route("/screenshots/<filename>")
@login_required
def screenshot_file(filename):
    filepath = os.path.join(SCREENSHOTS_DIR, filename)
    if os.path.exists(filepath):
        return send_from_directory(SCREENSHOTS_DIR, filename)
    return "", 404


@app.route("/api/export", methods=["POST"])
@login_required
def export_leads():
    import json as _json
    import io
    import openpyxl
    from flask import send_file
    ids = set(_json.loads(request.form.get("ids", "[]")))
    leads = [l for l in load_leads(current_user.id) if l["id"] in ids]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Naam", "Type", "Niche", "Stad", "Adres", "Website", "Email", "Telefoon", "Status", "Kwaliteit", "Notitie"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    quality_labels = {"q-low": "Weinig gevonden", "q-mid": "Gedeeltelijk", "q-high": "Goede lead", "q-full": "Volledig"}

    for l in leads:
        score = sum([bool(l.get("website")), bool(l.get("email")), bool(l.get("phone")), bool(l.get("board_page") or l.get("contact_person"))])
        qkey = "q-low" if score <= 1 else "q-mid" if score <= 2 else "q-high" if score <= 3 else "q-full"
        ws.append([
            l.get("name", ""),
            "Sportclub" if l.get("type") == "sport" else "Eetgelegenheid",
            l.get("niche", ""),
            l.get("city", ""),
            l.get("address", ""),
            l.get("website", ""),
            l.get("email", ""),
            l.get("phone", ""),
            l.get("status", "nieuw"),
            quality_labels.get(qkey, ""),
            l.get("note", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="leads_export.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Geen bestand"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        names = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.strip():
                    names.append(cell.strip())
        wb.close()
        return jsonify({"names": names})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


_CLUB_PREFIXES = re.compile(
    r"^(tennisvereniging|tennisclub|tennispark|tc |tv |sv |voetbalvereniging|voetbalclub|"
    r"hockeyvereniging|hockeyclub|handbalvereniging|basketbalvereniging|volleybalvereniging|"
    r"atletiekvereniging|zwemvereniging|golfclub|golfvereniging|padel|"
    r"sportvereniging|sportclub)\s*",
    re.IGNORECASE,
)

def _extract_city_from_name(name: str) -> str:
    """Probeer een plaatsnaam te extraheren uit een clubnaam, bijv. 'Tennisvereniging Albergen' → 'Albergen'."""
    stripped = _CLUB_PREFIXES.sub("", name).strip()
    # Verwijder lidwoorden zoals "De", "Het", "Den", "'t"
    stripped = re.sub(r"^(de|het|den|'t)\s+", "", stripped, flags=re.IGNORECASE).strip()
    words = stripped.split()
    # Neem aan dat het een stad is als het 1-3 woorden zijn die beginnen met hoofdletter
    if 1 <= len(words) <= 3 and all(w[0].isupper() for w in words if w):
        return stripped
    return ""


def run_manual_lead_job(job_id: str, name: str, force_type: str, user_id: int = None):
    """Zoek een club via Google Maps of DuckDuckGo. Voegt ALTIJD toe, ook zonder website."""
    jobs[job_id] = {"status": "running", "progress": 5, "message": f"Zoeken naar {name}...", "count": 0}
    _wait_until_slot_free(job_id)
    try:
        # Duplicate check — sla over als de lead al goede info heeft
        leads = load_leads(user_id)
        existing = next((l for l in leads if _names_similar(name, l["name"])), None)
        if existing and (existing.get("website") or existing.get("email") or existing.get("phone")):
            jobs[job_id] = {"status": "done", "progress": 100, "message": f"{name} — al in database", "count": 1}
            return

        extracted_city = _extract_city_from_name(name)
        result_name = name
        website, phone, address, city = "", "", "", extracted_city
        email, contact, logo_url, board_page = "", "", "", ""
        method = "handmatig"

        # Google Maps zoeken — meerdere queries proberen
        search_city = extracted_city or ""
        bare_name = _CLUB_PREFIXES.sub("", name).strip()
        bare_name_no_de = re.sub(r"^(de|het|den|'t)\s+", "", bare_name, flags=re.IGNORECASE).strip()
        queries_to_try = [
            (name, search_city),
            (bare_name, search_city),
            (bare_name_no_de, search_city),
            (name, ""),
            (bare_name, ""),
        ]
        first_result = None  # Beste fallback als naam niet matcht
        for q_name, q_city in queries_to_try:
            if website:
                break
            if not q_name.strip():
                continue
            try:
                jobs[job_id].update({"progress": 20, "message": f"Google Maps zoeken: {q_name}..."})
                businesses = _scrape_maps_with_watchdog(q_name, q_city, 5, timeout_s=150)
                match = next((b for b in businesses if _names_similar(name, b["name"], threshold=0.45)), None)
                if match:
                    result_name = match["name"]
                    website = match.get("website", "")
                    phone = match.get("phone", "")
                    address = match.get("address", "")
                    city = match.get("city", extracted_city)
                    method = "Google Maps"
                elif businesses and first_result is None:
                    first_result = businesses[0]  # Onthoud eerste resultaat als noodoptie
            except Exception:
                continue

        # Geen naam-match gevonden maar wel Google Maps resultaten — pak de beste kandidaat
        if not website and first_result:
            result_name = first_result["name"]
            website = first_result.get("website", "")
            phone = first_result.get("phone", "")
            address = first_result.get("address", "")
            city = first_result.get("city", extracted_city)
            method = "Google Maps (beste match)"

        # Fallback: URL-patronen raden als Google Maps niks vond
        if not website:
            try:
                jobs[job_id].update({"progress": 40, "message": f"Website raden voor {name}..."})
                website = guess_club_website(name)
                if website:
                    method = "URL-patroon"
            except Exception:
                pass

        # Stap 3: scrape email/logo/bestuur van website
        if website:
            try:
                jobs[job_id].update({"progress": 55, "message": "Email zoeken op website..."})
                email, contact = find_email_and_contact(website)
            except Exception:
                pass
            try:
                jobs[job_id]["progress"] = 70
                logo_url = find_logo(website)
            except Exception:
                pass
            if force_type == "sport":
                try:
                    board_page = find_board_page(website)
                except Exception:
                    pass

        # Altijd opslaan — ook als alles mislukt
        lead = {
            "id": str(uuid.uuid4()),
            "name": result_name,
            "type": force_type,
            "niche": name,
            "city": city,
            "address": address,
            "phone": phone,
            "website": website,
            "email": email,
            "contact_person": contact,
            "logo_url": logo_url,
            "board_page": board_page,
            "website_status": "",
            "screenshot": "",
            "cold_email": "",
            "status": "nieuw",
            "added": datetime.now().isoformat(),
        }

        try:
            lead["website_status"] = check_website(website) if website else ""
        except Exception:
            pass
        try:
            jobs[job_id]["progress"] = 85
            save_email(lead["id"], generate_cold_email(lead))
        except Exception:
            pass

        gevonden_info = [v for v in ["website" if website else "", "email" if email else "", "telefoon" if phone else ""] if v]
        details = ", ".join(gevonden_info) if gevonden_info else "weinig gevonden"

        add_lead(lead, user_id)
        jobs[job_id] = {
            "status": "done", "progress": 100,
            "message": f"{result_name} toegevoegd via {method} ({details})",
            "count": 1,
        }
    except Exception as e:
        log.exception("Handmatige lead-job mislukt voor %r", name)
        # Zet job altijd op done zodat de frontend niet blijft hangen
        jobs[job_id] = {"status": "done", "progress": 100, "message": f"{name} — fout: {e}", "count": 0}


@app.route("/api/leads/manual", methods=["POST"])
@login_required
def add_manual_lead():
    """Start een job die de club zoekt via DuckDuckGo en scrapet."""
    data = request.json or {}
    name = (data.get("name") or "").strip()
    force_type = data.get("force_type", "sport")
    if not name:
        return jsonify({"error": "Naam verplicht"}), 400

    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "running", "progress": 0, "message": "Starten...", "count": 0}
    _job_executor.submit(run_manual_lead_job, job_id, name, force_type, current_user.id)
    return jsonify({"job_id": job_id})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port, threaded=True)
